from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook
import utils.validaciones as utils_validacion
import utils.ventana as utils_ventana
import producto_lote.consultas as producto_lote_consultas
import PySimpleGUI as psg
import os
from datetime import datetime


from producto_lote.producto_lote import Producto_lote


def obtener_alias_controlador() -> str:
    return 'producto_lote_importar'


def procesar(
        ventana_actual=None,
        ventana_secundaria=None,
        ventana_auxiliar=None,
        cmd=None,
        valores=None,
        conn=None
):
    ventana = None
    fecha = datetime.today().isoformat(timespec='minutes')
    fecha = fecha.replace(':', '')
    campos = ['SKU', 'Nombre', 'Costo unitario', 'Porcentaje impuesto',
              'Monto impuesto', 'Monto utilidad', 'Precio', 'Redondeo',
              'Precio final', 'Cantidad', 'Unidad', 'Disponible',
              'Reservado', 'Estado', 'Días vida útil']

    if cmd['evt'] != obtener_alias_controlador():
        psg.popup_error('Intentando procesar un evento de producto, pero se envió otro evento',
                        'Error de programación')
    else:
        if cmd['det'] == 'plantilla':
            if cmd['act'] == 'procesar':
                destino_nombre = valores['destino-nombre']

                nombre = 'plantilla-productos-{}.xlsx'.format(fecha)
                nombre_archivo = os.path.join(destino_nombre, nombre)
                wb = Workbook()
                ws: Worksheet = wb.active
                ws.title = 'Lote de productos'
                for i, dato in enumerate(campos, 1):
                    ws.cell(row=1, column=i).value = dato.upper()
                    # @ es la letra antes de la A, para que cuando se sume 1, sea A la primera columna
                    ws.column_dimensions[chr(ord('@') + i)].width = len(campos[i - 1]) + 3

                wb.save(nombre_archivo)
                print("LA GATA ")
                ventana_actual['mensaje-error'].update('Archivo: {}'.format(nombre_archivo))
                ventana_actual.close()
                ventana = ventana_secundaria

        elif cmd['det'] == 'importar':
            if cmd['act'] == 'procesar':
                origen_nombre = valores['origen-nombre']

                wb = load_workbook(filename=origen_nombre)
                ws = wb.worksheets[0]

                todos_los_productos_lote_lote = producto_lote_consultas.listar_todos(conn=conn)
                todos_los_sku = set([p.sku for p in todos_los_productos_lote])
                print(todos_los_sku)

                for row in ws.iter_rows(min_row=2, max_col=len(campos), values_only=True):
                    sku, nombre, costo, porc_imp, monto_imp, monto_util, precio, redondeo, precio_final, cantidad, unidad, disponible, reservado, estado, vida_util = row
                    if sku not in todos_los_sku:
                        print(sku, nombre, costo, porc_imp, monto_imp, monto_util, precio, redondeo, precio_final,
                              cantidad, unidad, disponible, reservado, estado, vida_util)
                        producto_lote = Producto_lote()
                        producto_lote.sku = sku
                        producto_lote.nombre = nombre
                        producto_lote.costo_unitario = costo
                        producto_lote.porcentaje_impuesto = porc_imp
                        producto_lote.monto_impuesto = monto_imp
                        producto_lote.monto_utilidad = monto_util
                        producto_lote.precio = precio
                        producto_lote.redondeo = redondeo
                        producto_lote.precio_final = precio_final
                        producto_lote.cantidad = cantidad
                        producto_lote.unidad = unidad
                        producto_lote.disponible = disponible
                        producto_lote.reservado = reservado
                        producto_lote.estado = estado
                        producto_lote.dias_vida_util = vida_util
                        producto_lote_consultas.registrar(conn=conn, dato=producto_lote)
                    else:
                        print('Esta ', sku)
                # ventana_actual['mensaje-error'].update('Archivo: {}'.format(nombre_archivo))
                ventana_actual.close()
                ventana = ventana_secundaria
                # ventana_actual['mensaje-error'].update('Archivo: {}'.format(origen_nombre))
    return ventana
