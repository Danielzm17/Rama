import utils.validaciones as utils_validacion
import utils.ventana as utils_ventana
import producto.consultas as producto_consultas
import PySimpleGUI as psg
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

from producto.producto import Producto


def obtener_alias_controlador() -> str:
    return 'producto_reporte'


def procesar(
        ventana_actual=None,
        ventana_secundaria=None,
        ventana_auxiliar=None,
        cmd=None,
        valores=None,
        conn=None
):
    ventana = None
    fecha = datetime.today().isoformat(timespec='minutes')
    fecha = fecha.replace(':', '')
    campos = ['SKU', 'Nombre', 'Costo unitario', 'Porcentaje impuesto',
              'Monto impuesto', 'Monto utilidad', 'Precio', 'Redondeo',
              'Precio final', 'Cantidad', 'Unidad', 'Disponible',
              'Reservado', 'Estado', 'Días vida útil']

    if cmd['evt'] != obtener_alias_controlador():
        psg.popup_error('Intentando procesar un evento de producto, pero se envió otro evento', 'Error de programación')
    else:
        if cmd['det'] == 'generar':
            if cmd['act'] == 'generar':
                tab = valores['Tab']
                if tab == 'tab-rep01':
                    sku = valores['rep01-filtro-sku']
                    nombre = valores['rep01-filtro-nombre']
                    ordenar_sku = valores['rep01-ordenar-sku']
                    ordenar_nombre = valores['rep01-ordenar-nombre']
                    productos = producto_consultas.listar_por_sku_nombre(conn=conn, sku=sku, nombre=nombre)

                    destino_nombre = valores['destino-nombre']

                    nombre = 'reporte-productos-{}.xlsx'.format(fecha)
                    nombre_archivo = os.path.join(destino_nombre, nombre)


                    wb = Workbook()
                    ws: Worksheet = wb.active
                    ws.title = 'Productos'
                    for i, dato in enumerate(campos, 1):
                        ws.cell(row=1, column=i).value = dato.upper()
                        # @ es la letra antes de la A, para que cuando se sume 1, sea A la primera columna
                        ws.column_dimensions[chr(ord('@') + i)].width = len(campos[i - 1]) + 3

                    offset = 2
                    indice = 0

                    for i in range(len(productos)):
                        ws.cell(row=offset+i,column=1).value = productos[i].sku
                        ws.cell(row=offset + i, column=2).value = productos[i].nombre
                        ws.cell(row=offset + i, column=3).value = productos[i].costo_unitario
                        ws.cell(row=offset + i, column=4).value = productos[i].porcentaje_impuesto
                        ws.cell(row=offset + i, column=5).value = productos[i].monto_impuesto
                        ws.cell(row=offset + i, column=6).value = productos[i].monto_utilidad
                        ws.cell(row=offset + i, column=7).value = productos[i].precio
                        ws.cell(row=offset + i, column=8).value = productos[i].redondeo
                        ws.cell(row=offset + i, column=9).value = productos[i].precio_final
                        ws.cell(row=offset + i, column=10).value = productos[i].cantidad
                        ws.cell(row=offset + i, column=11).value = productos[i].unidad
                        ws.cell(row=offset + i, column=12).value = productos[i].disponible
                        ws.cell(row=offset + i, column=13).value = productos[i].reservado
                        ws.cell(row=offset + i, column=14).value = productos[i].estado
                        ws.cell(row=offset + i, column=8).value = productos[i].dias_vida_util

                    wb.save(nombre_archivo)

                ventana_actual.close()
                ventana = ventana_secundaria


                ventana_actual.close()
                ventana = ventana_secundaria
                # ventana_actual['mensaje-error'].update('Archivo: {}'.format(origen_nombre))
    return ventana
